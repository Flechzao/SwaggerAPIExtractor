import yaml
from openpyxl import Workbook


def extract_api_info_from_yaml_file(yaml_file_path):
    with open(yaml_file_path, 'r') as file:
        swagger_yaml = yaml.safe_load(file)

    # 创建一个新的工作簿，并添加一个工作表
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "APIs"

    # 添加表头
    headers = ["Path", "Method", "Operation ID"]
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)

    row_num = 2
    for path, path_item in swagger_yaml['paths'].items():
        for method, operation in path_item.items():
            if method in ['get', 'post', 'put', 'delete', 'patch']:
                path_cell = worksheet.cell(row=row_num, column=1, value=path)
                method_cell = worksheet.cell(row=row_num, column=2, value=method.upper())
                operation_id_cell = worksheet.cell(row=row_num, column=3, value=operation.get('operationId', 'N/A'))
                row_num += 1

    # 保存工作簿到Excel文件
    excel_file_path = "output.xlsx"  # 设置输出Excel文件的路径
    workbook.save(excel_file_path)
    print(f"API信息已保存到 {excel_file_path}")


if __name__ == "__main__":
    yaml_file_path = "path/to/your/swagger.yaml"  # 替换为你的Swagger YAML文件的路径
    extract_api_info_from_yaml_file(yaml_file_path)
